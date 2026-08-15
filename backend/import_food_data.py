from pathlib import Path

import openpyxl

from database import Base, SessionLocal, engine
from models import Food, Protein, Vegetable

EXCEL_PATH = Path(__file__).resolve().parent / "food_data" / "FoodList.xlsx"

MAIN_PROTEIN_SHEETS = ["Chicken", "Beef", "Fish", "Turkey", "Egg", "Cheese"]

BREAKFAST_COLUMN_TO_PROTEIN = {
    "contains chicken": "Chicken",
    "contains beef": "Beef",
    "contains fish": "Fish",
    "contains turkey": "Turkey",
    "egg": "Egg",
    "cheese": "Cheese",
}


def normalize_header(value) -> str:
    return " ".join(str(value).split()).lower()


def clean_name(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def find_sheet(workbook, wanted_name):
    wanted = wanted_name.strip().lower()
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().lower() == wanted:
            return workbook[sheet_name]
    raise ValueError(f"Sheet '{wanted_name}' not found. Available sheets: {workbook.sheetnames}")


def food_already_exists(db, name, protein_id, meal_type) -> bool:
    return (
        db.query(Food)
        .filter(Food.name == name, Food.protein_id == protein_id, Food.meal_type == meal_type)
        .first()
        is not None
    )


def import_vegetables(db, workbook, report):
    sheet = find_sheet(workbook, "Vegetable")
    for row in sheet.iter_rows(min_row=2, values_only=True): 
        name = clean_name(row[0])
        if not name:
            continue
        if db.query(Vegetable).filter(Vegetable.name == name).first():
            report["duplicate_vegetables_skipped"].append(name)
            continue
        db.add(Vegetable(name=name))
        db.flush() 
        report["vegetables_inserted"] += 1
    db.commit()


def import_main_foods(db, workbook, report):
    for sheet_name in MAIN_PROTEIN_SHEETS:
        sheet = find_sheet(workbook, sheet_name)
        protein = db.query(Protein).filter(Protein.name == sheet_name).first()
        if not protein:
            report["unknown_proteins"].append(sheet_name)
            continue

        for row in sheet.iter_rows(min_row=2, values_only=True): 
            raw_name = row[0]
            name = clean_name(raw_name)
            if not name:
                continue
            if str(raw_name) != name:
                report["possible_typos_or_whitespace"].append(f"{sheet_name} sheet: {raw_name!r} -> {name!r}")

            if food_already_exists(db, name, protein.id, "main"):
                report["duplicate_foods_skipped"].append(f"{name} ({sheet_name}, main)")
                continue

            db.add(Food(name=name, protein_id=protein.id, meal_type="main"))
            db.flush()
            report["main_foods_inserted"] += 1
    db.commit()


def import_breakfast_foods(db, workbook, report):
    sheet = find_sheet(workbook, "Breakfast")
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_header(h) for h in rows[0]]
    protein_columns = [
        (col_index, BREAKFAST_COLUMN_TO_PROTEIN[header_value])
        for col_index, header_value in enumerate(header)
        if header_value in BREAKFAST_COLUMN_TO_PROTEIN
    ]

    for row in rows[1:]:
        name = clean_name(row[0])
        if not name:
            continue

        matched_proteins = [protein_name for col_index, protein_name in protein_columns if row[col_index] is True]

        if len(matched_proteins) != 1:
            report["breakfast_protein_problems"].append(f"{name}: {len(matched_proteins)} protein flags set, expected 1")
            continue

        protein = db.query(Protein).filter(Protein.name == matched_proteins[0]).first()
        if not protein:
            report["unknown_proteins"].append(matched_proteins[0])
            continue

        if food_already_exists(db, name, protein.id, "breakfast"):
            report["duplicate_foods_skipped"].append(f"{name} ({matched_proteins[0]}, breakfast)")
            continue

        db.add(Food(name=name, protein_id=protein.id, meal_type="breakfast"))
        db.flush()
        report["breakfast_foods_inserted"] += 1
    db.commit()


def print_report(report):
    print("\n=== Import report ===")
    print(f"Vegetables inserted:      {report['vegetables_inserted']}")
    print(f"Main foods inserted:      {report['main_foods_inserted']}")
    print(f"Breakfast foods inserted: {report['breakfast_foods_inserted']}")

    print(f"\nDuplicate vegetables skipped ({len(report['duplicate_vegetables_skipped'])}): "
          f"{report['duplicate_vegetables_skipped']}")
    print(f"Duplicate foods skipped ({len(report['duplicate_foods_skipped'])}): "
          f"{report['duplicate_foods_skipped']}")
    print(f"Unknown proteins found in the file ({len(report['unknown_proteins'])}): "
          f"{report['unknown_proteins']}")
    print(f"Breakfast rows skipped for protein-count problems ({len(report['breakfast_protein_problems'])}): "
          f"{report['breakfast_protein_problems']}")
    print(f"Possible typos / whitespace noticed, NOT auto-corrected ({len(report['possible_typos_or_whitespace'])}):")
    for item in report["possible_typos_or_whitespace"]:
        print(f"  - {item}")


if __name__ == "__main__":
    if not EXCEL_PATH.exists():
        raise SystemExit(f"Excel file not found: {EXCEL_PATH}")

    Base.metadata.create_all(bind=engine)
    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Loaded {EXCEL_PATH.name}. Sheets found: {workbook.sheetnames}")

    report = {
        "vegetables_inserted": 0,
        "main_foods_inserted": 0,
        "breakfast_foods_inserted": 0,
        "duplicate_vegetables_skipped": [],
        "duplicate_foods_skipped": [],
        "unknown_proteins": [],
        "breakfast_protein_problems": [],
        "possible_typos_or_whitespace": [],
    }

    db = SessionLocal()
    try:
        import_vegetables(db, workbook, report)
        import_main_foods(db, workbook, report)
        import_breakfast_foods(db, workbook, report)
    finally:
        db.close()

    print_report(report)
